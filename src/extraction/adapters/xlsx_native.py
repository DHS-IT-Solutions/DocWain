"""Native XLSX adapter using openpyxl.

Preserves every sheet (including hidden, flagged not dropped), every cell with
both value and formula where present, merged cells, named ranges.

Spec: docs/superpowers/specs/2026-04-23-extraction-accuracy-design.md §4.1
"""
from __future__ import annotations

import io

import openpyxl

from src.extraction.adapters.errors import NativeAdapterError
from src.extraction.canonical_schema import ExtractionResult, Sheet


def _cell_to_dict(cell) -> dict:
    value = cell.value
    data_type = cell.data_type or "n"
    formula = None
    if isinstance(value, str) and value.startswith("="):
        formula = value
    return {"value": value, "formula": formula, "type": data_type}


def extract_xlsx_native(file_bytes: bytes, *, doc_id: str, filename: str) -> ExtractionResult:
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=False)
    except Exception as exc:
        raise NativeAdapterError(f"failed to open XLSX {filename!r}: {exc}") from exc

    sheets: list[Sheet] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        cells = {}
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                cells[(cell.row, cell.column)] = _cell_to_dict(cell)

        merged = [str(rng) for rng in ws.merged_cells.ranges]
        named = [n for n in wb.defined_names if wb.defined_names[n].attr_text and sheet_name in (wb.defined_names[n].attr_text or "")]

        sheets.append(
            Sheet(
                name=sheet_name,
                cells=cells,
                hidden=(ws.sheet_state != "visible"),
                merged_cells=merged,
                named_ranges=named,
            )
        )

    return ExtractionResult(
        doc_id=doc_id,
        format="xlsx",
        path_taken="native",
        sheets=sheets,
    )
