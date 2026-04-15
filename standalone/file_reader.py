import base64
import csv
import io
from pathlib import Path


class UnsupportedFileType(Exception):
    pass


_EXT_MAP = {
    ".pdf": "pdf",
    ".docx": "docx",
    ".doc": "docx",
    ".xlsx": "xlsx",
    ".xls": "xlsx",
    ".csv": "csv",
    ".tsv": "tsv",
    ".txt": "text",
    ".text": "text",
    ".md": "text",
    ".json": "text",
    ".xml": "text",
    ".html": "text",
    ".htm": "text",
    ".png": "image",
    ".jpg": "image",
    ".jpeg": "image",
    ".tiff": "image",
    ".tif": "image",
    ".bmp": "image",
    ".webp": "image",
}

_SUPPORTED_TYPES = {"pdf", "docx", "xlsx", "csv", "tsv", "text", "image"}


def detect_file_type(filename: str, data: bytes) -> str:
    ext = Path(filename).suffix.lower()
    if ext in _EXT_MAP:
        return _EXT_MAP[ext]
    if data[:5] == b"%PDF-":
        return "pdf"
    if data[:4] == b"PK\x03\x04":
        return "docx"  # could be docx or xlsx, default docx
    return "unknown"


def read_file(filename: str, data: bytes) -> str:
    file_type = detect_file_type(filename, data)

    if file_type not in _SUPPORTED_TYPES:
        raise UnsupportedFileType(f"Unsupported file type: {filename}")

    if file_type == "text":
        return data.decode("utf-8", errors="replace")

    if file_type == "csv":
        return _read_csv(data)

    if file_type == "tsv":
        return _read_csv(data, delimiter="\t")

    if file_type == "pdf":
        return _read_pdf(data)

    if file_type == "docx":
        return _read_docx(data)

    if file_type == "xlsx":
        return _read_xlsx(data)

    if file_type == "image":
        return _read_image(filename, data)

    raise UnsupportedFileType(f"Unsupported file type: {filename}")


def read_file_with_metadata(filename: str, data: bytes) -> tuple[str, dict]:
    file_type = detect_file_type(filename, data)
    content = read_file(filename, data)
    meta = {
        "file_type": file_type,
        "filename": filename,
        "size_bytes": len(data),
    }
    if file_type == "pdf":
        meta["pages"] = _count_pdf_pages(data)
    elif file_type == "image":
        meta["pages"] = 1
    return content, meta


def _read_csv(data: bytes, delimiter: str = ",") -> str:
    text = data.decode("utf-8", errors="replace")
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    rows = list(reader)
    if not rows:
        return ""
    lines = []
    for row in rows:
        lines.append(" | ".join(row))
    return "\n".join(lines)


def _read_pdf(data: bytes) -> str:
    try:
        import fitz  # PyMuPDF

        doc = fitz.open(stream=data, filetype="pdf")
        pages = []
        for page in doc:
            pages.append(page.get_text())
        doc.close()
        return "\n\n".join(pages)
    except ImportError:
        raise ImportError(
            "PyMuPDF (fitz) is required for PDF processing. Install with: pip install PyMuPDF"
        )


def _read_docx(data: bytes) -> str:
    try:
        from docx import Document

        doc = Document(io.BytesIO(data))
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
        return "\n\n".join(paragraphs)
    except ImportError:
        raise ImportError(
            "python-docx is required for DOCX processing. Install with: pip install python-docx"
        )


def _read_xlsx(data: bytes) -> str:
    try:
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        sheets = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                rows.append(" | ".join(cells))
            if rows:
                sheets.append(f"Sheet: {sheet_name}\n" + "\n".join(rows))
        wb.close()
        return "\n\n".join(sheets)
    except ImportError:
        raise ImportError(
            "openpyxl is required for Excel processing. Install with: pip install openpyxl"
        )


def _read_image(filename: str, data: bytes) -> str:
    ext = Path(filename).suffix.lower().lstrip(".")
    mime_map = {
        "png": "image/png",
        "jpg": "image/jpeg",
        "jpeg": "image/jpeg",
        "tiff": "image/tiff",
        "tif": "image/tiff",
        "bmp": "image/bmp",
        "webp": "image/webp",
    }
    mime = mime_map.get(ext, "image/png")
    b64 = base64.b64encode(data).decode("ascii")
    return f"data:{mime};base64,{b64}"


def _count_pdf_pages(data: bytes) -> int:
    try:
        import fitz
        doc = fitz.open(stream=data, filetype="pdf")
        count = len(doc)
        doc.close()
        return count
    except (ImportError, Exception):
        return 1
