import io

import openpyxl

from src.extraction.adapters.xlsx_native import extract_xlsx_native
from src.extraction.canonical_schema import ExtractionResult


def _make_xlsx(sheets: dict[str, list[list]]) -> bytes:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for name, rows in sheets.items():
        ws = wb.create_sheet(title=name)
        for r_idx, row in enumerate(rows, start=1):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_xlsx_native_extracts_single_sheet():
    xlsx = _make_xlsx({"Sheet1": [["h1", "h2"], [1, 2], [3, 4]]})
    result = extract_xlsx_native(xlsx, doc_id="d1", filename="t.xlsx")
    assert isinstance(result, ExtractionResult)
    assert result.format == "xlsx"
    assert len(result.sheets) == 1
    assert result.sheets[0].name == "Sheet1"
    assert result.sheets[0].cells[(1, 1)]["value"] == "h1"
    assert result.sheets[0].cells[(2, 1)]["value"] == 1


def test_xlsx_native_preserves_multiple_sheets():
    xlsx = _make_xlsx({"A": [["x"]], "B": [["y"]]})
    result = extract_xlsx_native(xlsx, doc_id="d2", filename="t.xlsx")
    names = [s.name for s in result.sheets]
    assert names == ["A", "B"]


def test_xlsx_native_flags_hidden_sheet():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws1 = wb.create_sheet(title="Visible")
    ws1["A1"] = "v"
    ws2 = wb.create_sheet(title="Hidden")
    ws2["A1"] = "h"
    ws2.sheet_state = "hidden"
    buf = io.BytesIO()
    wb.save(buf)

    result = extract_xlsx_native(buf.getvalue(), doc_id="d3", filename="t.xlsx")
    hidden = [s for s in result.sheets if s.hidden]
    assert len(hidden) == 1
    assert hidden[0].name == "Hidden"
